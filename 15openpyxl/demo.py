"""
excel表格需要为.xlsx格式
三大对象：
    workbook:工作簿对象
    sheet:表单对象
    cell:表格对象

"""
import openpyxl

# -------------------------------基础用法-------------------------------------
# 加载excel文件为工作簿对象
workbook = openpyxl.load_workbook(r'E:\PythonBase\15openpyxl\demo1.xlsx')
# 获取所有的表单
print(workbook.sheetnames)
# 选中表单
sh = workbook['Sheet1']
# 读取数据
c = sh.cell(row=1, column=1)
print(c.value)

# -------------------------------读取表单中所有的数据-------------------------------------
workbook = openpyxl.load_workbook(r'E:\PythonBase\15openpyxl\demo1.xlsx')
sh = workbook['Sheet1']
#  rows:按行获取表单中所有的格子，以列表嵌套元组返回,每一行的格子放到一个元组中
res = list(sh.rows)
print(res)
for i in res:
    print(i)

# columns:按列读取表单中的所有格子，以列表嵌套元组返回，每一列的格子放到一个元组中
res = list(sh.columns)
for i in res:
    print(i)