import openpyxl


class HandleExcel():
    def __init__(self, filename, sheetname):
        """

        :param filename: excle文件名（路径）
        :param sheetname: 表单
        """
        self.filename = filename
        self.sheetname = sheetname

    def read_excel(self):
        """
        读取excel数据
        :param filename: excle文件名（路径）
        :param sheetname: 表单
        :return:
        """

        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheetname]

        res = list(sh.rows)
        # 获取第一行的表头
        title = [i.value for i in res[0]]
        cases = []
        # 遍历第一行以外的其他行
        for item in res[1:]:
            data = [i.value for i in item]
            dic = dict(zip(title, data))
            cases.append(dic)
        # 返回读取出来的数据
        return cases

    def write_excel(self, row, column, value):
        """
        数据写入的方法
        :param row:写入的行
        :param column:写入的列
        :param value:写入的值
        :return:
        """
        wb = openpyxl.load_workbook(self.filename)
        sh = wb[self.sheetname]

        # 写入数据
        sh.cell(row=row, column=column, value=value)
        # 保存表格内容
        wb.save(self.filename)
